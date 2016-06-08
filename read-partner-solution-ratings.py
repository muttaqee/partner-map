#!/usr/bin/env python

"""
Read data for partner_strengths_ratings table:
- Read rows from "Map" worksheet.
- Encode rows in JSON. In this form:
  {
    ...,
    "Piper Enterprise Solutions": {
        "Financial Rate Negotiation": "B",
        "Political - SAS/Customer": "B",
        "Process & Training": "B",
        "Social - Responsive": "B",
        "Technical - Quality": "C"
    },
    ...
  }
- Output rows for use by PHP script.

06-06-2016

FIXME: Update header comments here
"""

import openpyxl, json
from openpyxl.cell import column_index_from_string, get_column_letter

# Command-line arguments # FIXME: Use this if script is general-purpose
# [0] File name
# [1] Workbook name
# [2] Worksheet name
# [3] col_first
# [4] row_first
# [5] column names list
# [6] Number of rows

#def connect():
# Open file
wb_name = "partner-spreadsheet-copy.xlsx"
try:
    wb = openpyxl.load_workbook(wb_name, data_only = True)
except:
    print("Error: Could not load workbook.")

# Official_Names sheet
ws_name = "Map"
try:
    ws = wb.get_sheet_by_name(ws_name)
except:
    print("Error: Could not load worksheet.")

# Columns and rows in worksheet to remember
super_columns = ["CFS", "CIS", "PMS", "SCS", "HLS", "RTS", "EN"]
columns = [
"Fraud and Financial Crimes",
"Anti-Money Laundering",
"Credit Scoring", # 3 (4) NOTE: The excel sheet has a duplicate column for this.
"Credit Risk Managment",
"Risk Dimensions / Management",
"OpRisk Management",
"Enterprise GRC",
"CFS Solutions", # 8 (9)

"Marketing Automation",
"Marketing Optimization",
"Rel-Time Decision Mgr",
"Marketing Operations Management",
"Realtime Decision Manager",
"CI Solutions", # 14 (15)

"ABM / Profictability Managament",
"Strategy Management",
"Financial Management",
"Human Capital Management",
"PM Solutions", # 19 (20)

"Collaborative Planning Workbench",
"Demand Signal Analytics",
"Forecast Analyst Workbench",
"New Product Forecasting",
"Asset Performance Analytics",
"Field Quality Analytics",
"Production Quality Analytics",
"Suspect Claims Detection",
"Service Parts Optimization",
"Inventory Optimization",
"SC Solutions", # 30 (31)

"Clinical Data Integration",
"Drug Development",
"Healthcare Fraud",
"Episode Analytics",
"Safety Analytics",
"Claims Analytics",
"Health Life Sci Solutions", # 37 (38)

"Integrated Merchandise Planning",
"Revenue Optimization",
"Size/Pack Optimization",
"Demand-Driven Forecasting",
"Retail Solutions", # 42 (43)

"Energy Forecasting" # 43 (44)
]
x_first = column_index_from_string('AT')
x_last = column_index_from_string('CK')
y_first = 4
y_last = 106
width = len(columns)
height = y_last - y_first + 1 # Number of rows

# Create JSON name-value pair. NOTE: Only name parameter is encoded as string
def getNameValuePair(name, value):
    return "\"" + name + "\"" + ":" + value

def getSuperColNameFromX(x):
    idx = 0;
    if x < x_first + 9:
        idx = 0
    elif x < x_first + 15:
        idx = 1
    elif x < x_first + 20:
        idx = 2
    elif x < x_first + 31:
        idx = 3
    elif x < x_first + 38:
        idx = 4
    elif x < x_first + 43:
        idx = 5
    elif x < e_first + 44:
        idx = 6
    return super_columns[idx]

def getColNameFromX(x):
    return columns[x - column_index_from_string("AT")] # FIXME: is 14 correct?

def acceptString(s):
    acceptable_strings = [
    "A", "B", "C", "D", "E", "F"
    ]
    for string in acceptable_strings:
        if s == string:
            return True;
    return False;

# Read in and associate data with partner_name, technology, or rating db columns
def encodeRow(x, y, rating_value):
    json_row = ""

    partner_name = str.strip(ws.cell(row = y, column = 2).value)
    solution_type = getSuperColNameFromX(x)
    solution = getColNameFromX(x)
    rating = rating_value
    #print("     ", getSuperColNameFromX(x), getColNameFromX(x))

    json_row += getNameValuePair("partner_name", "\"" + partner_name + "\"") + ","
    json_row += getNameValuePair("solution_type", "\"" + solution_type + "\"") + ","
    json_row += getNameValuePair("solution", "\"" + solution + "\"") + ","
    json_row += getNameValuePair("rating", "\"" + rating + "\"")

    return "{" + json_row + "}"

def encodeAllRows(x0, x1, y0, y1):
    json_collection = ""

    id = 1
    for y in range (y0, y1+1):
        for x in range (x0, x1+1):
            cell_value = (str.strip(str(ws.cell(row = y, column = x).value))).upper()
            if cell_value and acceptString(cell_value):
                json_collection += getNameValuePair(str(id), encodeRow(x, y, rating_value = cell_value)) + ","
                id += 1

    return "{" + json_collection[:-1] + "}"

# Read data from spreadsheet, convert to JSON, and output:

# FIXME - remove this try/except
# try:
#     connect();
# except:
#     print("Error: Could not open workbook or worksheet.")
try:
    parsed_json = json.loads(encodeAllRows(x0 = x_first, x1 = x_last, y0 = y_first, y1 = y_last))
except:
    print("Error: Could not load data from spreadsheet.")
try:
    print(json.dumps(parsed_json))
    # Human-readable version:
    # print(json.dumps(parsed_json, indent = 4, sort_keys = True))
except:
    print("Error: Could not output spreadsheet data in JSON.")
