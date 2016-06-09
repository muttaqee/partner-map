#!/usr/bin/env python

"""
Read data for partner_vertical_junction table:
- Read rows from "Map" worksheet.
- Encode rows in JSON. In this form:
  {
    ...,
    "1": {
        "partner_name":"ANALYTIKS INTERNATIONAL INC",
        "misc_type":"Platform Administration",
        "rating":"A"
    },
    ...
  }
- Output rows for use by PHP script.

06-09-2016

FIXME: Update header comments here
"""

import openpyxl, json
from openpyxl.cell import column_index_from_string, get_column_letter

# Command-line arguments # FIXME: Use this if script is general-purpose
# [0] File name
# [1] Workbook name
# [2] Worksheet name
# [3] col_first
# [4] row_first
# [5] column names list
# [6] Number of rows

#def connect():
# Open file
wb_name = "partner-spreadsheet-copy.xlsx"
try:
    wb = openpyxl.load_workbook(wb_name, data_only = True)
except:
    print("Error: Could not load workbook.")

# Official_Names sheet
ws_name = "Consultants"
try:
    ws = wb.get_sheet_by_name(ws_name)
except:
    print("Error: Could not load worksheet.")

# Columns and rows in worksheet to remember
y_first = 3
y_last = 479

# Create JSON name-value pair. NOTE: Only name parameter is encoded as string
def getNameValuePair(name, value):
    return "\"" + name + "\"" + ":" + value

# Read in and associate data with partner_name, technology, or rating db columns
def encodeRow(y):
    json_row = ""

    last_name = str.strip(ws.cell(row = y, column = 1).value)
    tmp = ws.cell(row = y, column = 2).value
    if tmp:
        rating = tmp
    else:
        rating = "null"

    json_row += getNameValuePair("last_name", "\"" + last_name + "\"") + ","
    json_row += getNameValuePair("rating", "\"" + rating + "\"") + ","
    json_row += getNameValuePair("is_rejected", str(0))

    return "{" + json_row + "}"

def encodeAllRows(y0, y1):
    json_collection = ""

    id = 1
    for y in range (y0, y1+1):
        json_collection += getNameValuePair(str(id), encodeRow(y)) + ","
        id += 1

    return "{" + json_collection[:-1] + "}"

# Read data from spreadsheet, convert to JSON, and output:

# FIXME - remove this try/except
# try:
#     connect();
# except:
#     print("Error: Could not open workbook or worksheet.")
try:
    parsed_json = json.loads(encodeAllRows(y0 = y_first, y1 = y_last))
except:
    print("Error: Could not load data from spreadsheet.")
try:
    #print(json.dumps(parsed_json))
    # Human-readable version:
    print(json.dumps(parsed_json, indent = 4, sort_keys = True))
except:
    print("Error: Could not output spreadsheet data in JSON.")
