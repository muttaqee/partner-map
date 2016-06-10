#!/usr/bin/env python

"""
Read data for consultant_ratings table:
- Read rows from "Map" worksheet.
- Encode rows in JSON. In this form:
  {
    ...,
    "1": {
        "last_name":"Matt Ware",
        "area":"admin",
        "rating":"A"
    },
    ...
  }
- Output rows for use by PHP script.

06-10-2016

FIXME: Update header comments here
"""

import openpyxl, json
from openpyxl.cell import column_index_from_string, get_column_letter

# Command-line arguments # FIXME: Use this if script is general-purpose
# [0] File name
# [1] Workbook name
# [2] Worksheet name
# [3] col_first
# [4] row_first
# [5] column names list
# [6] Number of rows

#def connect():
# Open file
wb_name = "partner-spreadsheet-copy.xlsx"
try:
    wb = openpyxl.load_workbook(wb_name, data_only = True)
except:
    print("Error: Could not load workbook.")

# Official_Names sheet
ws_name = "Consultants"
try:
    ws = wb.get_sheet_by_name(ws_name)
except:
    print("Error: Could not load worksheet.")

# Columns and rows in worksheet to remember
columns = [
"programmer", "DI", "BI", "admin", "grid", "VA", "analytics"
]
x_first = column_index_from_string('D')
x_last = column_index_from_string('J')
y_first = 3
y_last = 479
width = len(columns)
height = y_last - y_first + 1 # Number of rows

# Create JSON name-value pair. NOTE: Only name parameter is encoded as string
def getNameValuePair(name, value):
    return "\"" + name + "\"" + ":" + value

def getColNameFromX(x):
    return columns[x - x_first]

def acceptString(s):
    acceptable_strings = [
    "A", "B", "C", "D", "F"
    ]
    for string in acceptable_strings:
        if s == string:
            return True;
    return False;

# Read in and associate data with partner_name, technology, or rating db columns
def encodeRow(x, y, rating_value):
    json_row = ""

    last_name = str.strip(ws.cell(row = y, column = 1).value)
    area = getColNameFromX(x)
    rating = rating_value

    json_row += getNameValuePair("last_name", "\"" + last_name + "\"") + ","
    json_row += getNameValuePair("area", "\"" + area + "\"") + ","
    json_row += getNameValuePair("rating", "\"" + rating + "\"")

    return "{" + json_row + "}"

def encodeAllRows(x0, x1, y0, y1):
    json_collection = ""

    id = 1
    for y in range (y0, y1+1):
        for x in range (x0, x1+1):
            cell_value = (str.strip(str(ws.cell(row = y, column = x).value))).upper()
            if cell_value and acceptString(cell_value):
                #json_collection += getNameValuePair(str(id), "\"value\"") + ","
                json_collection += getNameValuePair(str(id), encodeRow(x, y, cell_value)) + ","
                id += 1

    return "{" + json_collection[:-1] + "}"

# Read data from spreadsheet, convert to JSON, and output:

# FIXME - remove this try/except
# try:
#     connect();
# except:
#     print("Error: Could not open workbook or worksheet.")
try:
    parsed_json = json.loads(encodeAllRows(x0 = x_first, x1 = x_last, y0 = y_first, y1 = y_last))
except:
    print("Error: Could not load data from spreadsheet.")
try:
    print(json.dumps(parsed_json))
    # Human-readable version:
    # print(json.dumps(parsed_json, indent = 4, sort_keys = True))
except:
    print("Error: Could not output spreadsheet data in JSON.")
