#!/usr/bin/env python

"""
Read data for PARTNER table, pass 1/2:
- Read rows from "Map" worksheet.
- Encode rows in JSON.
- Output rows for use by PHP script.
"""

"""
NOTE to self:

Refer to...
read-save-xls-file.py --- to open and read from partner-spreadsheet-copy.xls
io-json-send.py --- to encode data in JSON and output

then refer to...
io-json-receive.php --- to receive JSON as associative array

and store data into tables (write and run table-filling script after running
wipe-init-db.php script)

5/26/2016

FIXME: Update header comments here
"""

import openpyxl, json
from openpyxl.cell import column_index_from_string

# Open file
wb_name = "partner-spreadsheet-copy.xlsx"
wb = openpyxl.load_workbook(wb_name)

# Official_Names sheet
ws_name = "Map"
ws = wb.get_sheet_by_name(ws_name)

# Column and row numbers to remember
x_begin = 2
x_end = column_index_from_string('DI') # FIXME: Need DI or DJ
x_name = 2 # Official name column
x_pp = 3 # Partner plus column
x_notes = x_end - 1; # Notes column

y_begin = 4
y_end = 107
count = y_end - y_begin # Number of rows

# Create JSON name-value pair. NOTE: Only name parameter is encoded as string
def getNameValuePair(name, value):
    return "\"" + name + "\"" + ":" + value

# For a given row, build and return an object in JSON
def getRow(row_number):
    json_row = ""
    if ws.cell(row = row_number, column = x_name).value:
        # "Unofficial" name (always null here) - FIXME: May remove; in PHP,
        # specifiy which columns to enter data into and leave this one out
        cell_value = str.strip(ws.cell(row = row_number, column = x_name).value);
        json_row += getNameValuePair(name = "name", value = "null") + ","

        # Official name (never null)
        json_row += getNameValuePair(name = "official_name", value = "\"" + cell_value + "\"") + ","

        # Partner plus (1 or 0)
        if ws.cell(row = row_number, column = x_pp).value:
            json_row += getNameValuePair(name = "is_partner_plus", value = str(1)) + ","
        else:
            json_row += getNameValuePair(name = "is_partner_plus", value = str(0)) + ","

        # Notes (string or null)
        if ws.cell(row = row_number, column = x_notes).value:
            cell_value = str.strip(str(ws.cell(row = row_number, column = x_notes).value))
            json_row += getNameValuePair(name = "notes", value = "\"" + cell_value + "\"")
        else:
            json_row += getNameValuePair(name = "notes", value = "null")

    return "{" + json_row + "}"

# Read data from each row - watch for null cells
def getAllRows(begin, end):
    json_collection = ""
    id = 1
    for y in range(begin, end-1):
        json_collection += getNameValuePair(name = str(id), value = getRow(y)) + ","
        id += 1
    # Last name-value pair has no comma:
    json_collection += getNameValuePair(name = str(id), value = getRow(end-1))
    return "{" + json_collection + "}"

# Read data from spreadsheet, convert to JSON, and output:
try:
    parsed_json = json.loads(getAllRows(y_begin, y_end))
except:
    print("Error: Could not load data from spreadsheet.")
try:
    print(json.dumps(parsed_json))
    # Human-readable version:
    # print(json.dumps(parsed_json, indent=4, sort_keys=True))
except:
    print("Error: Could not output spreadsheet data in JSON.")
